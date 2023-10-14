import os
import smtplib
import email
import collections
from email.mime.base import MIMEBase
from email import encoders
import imapclient
import datetime
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from tkinter import ttk, messagebox
from tkinter.tix import Balloon
import openpyxl
from tkcalendar import DateEntry
import tkinter.messagebox
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pdf2image import convert_from_path
from PIL import Image, ImageTk
import PyPDF2
import docx
import re
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter import filedialog, Listbox, EXTENDED
from tkinter import Tk, Entry, Button
from cryptography.fernet import Fernet
import shutil

root = Tk()

# Generate a key and save it - this would be done once
key = Fernet.generate_key()
with open("key.key", "wb") as key_file:
    key_file.write(key)

# Encrypt the password and save it - this would also be done once
cipher_suite = Fernet(key)
password = "ltqftpkngidjthlh"  # Your password here
encrypted_text = cipher_suite.encrypt(password.encode())  # Must be bytes
with open("encrypted.txt", "wb") as file:
    file.write(encrypted_text)

PIN_CODE = "2898"  # Replace 1234 with your chosen pin code


def check_pin_code():
    entered_pin = pin_entry.get()  # Get the pin from your entry widget
    if entered_pin == PIN_CODE:
        with open("key.key", "rb") as key_file:
            key = key_file.read()
        cipher_suite = Fernet(key)
        with open("encrypted.txt", "rb") as file:
            encrypted_text = file.read()
        password = cipher_suite.decrypt(
            encrypted_text
        ).decode()  # Decrypt and decode to string
        password_entry.delete(0, tk.END)  # Clear the password_entry
        password_entry.insert(0, password)  # Insert the decrypted password
    else:
        print("Invalid pin code")


pin_entry = Entry(root)
pin_entry.pack()

check_button = Button(root, text="Check pin", command=check_pin_code)
check_button.pack()


frame_index = 0
global results_listbox

downloaded_files_list = []

files_listbox = None  # Global listbox variable
pdf_window = None


def close_application():
    root.destroy()


def delete_directory():
    directory_path = "C:\\Users\\markz\\OneDrive\\Documenten\\downloaded pdfs"
    try:
        shutil.rmtree(directory_path)
        # After deleting, recreate the directory for future use
        os.makedirs(directory_path, exist_ok=True)

        # Clear the Listbox
        files_listbox.delete(0, tk.END)

        tk.messagebox.showinfo(
            "Success", "Directory and its contents have been deleted successfully!"
        )
    except Exception as e:
        tk.messagebox.showerror(
            "Error", f"Failed to delete the directory. Reason: {str(e)}"
        )


def load_output_folder():
    try:
        with open("output_folder.txt", "r") as f:
            return f.read().strip()
    except FileNotFoundError:
        return ""


def save_output_folder(folder_path):
    with open("output_folder.txt", "w") as f:
        f.write(folder_path)


unique_keywords = set()


def load_keywords():
    global unique_keywords
    try:
        with open("keywords.txt", "r") as f:
            unique_keywords = {line.strip() for line in f.readlines()}
            return list(unique_keywords)
    except FileNotFoundError:
        return []


def save_keyword(keyword):
    global unique_keywords
    unique_keywords.add(keyword)
    with open("keywords.txt", "w") as f:
        for kw in unique_keywords:
            f.write(f"{kw}\n")


def remove_keyword():
    global unique_keywords
    selected_keyword = search_keyword_entry.get()
    if selected_keyword in unique_keywords:
        unique_keywords.remove(selected_keyword)
        with open("keywords.txt", "w") as f:
            for kw in unique_keywords:
                f.write(f"{kw}\n")
        search_keyword_entry["values"] = list(unique_keywords)
        if unique_keywords:
            search_keyword_entry.set(list(unique_keywords)[-1])
        else:
            search_keyword_entry.set("")


def clean_file_name(file_name):
    invalid_chars = ["\\", "/", ":", "*", "?", '"', "<", ">", "|", "\r", "\n"]
    for char in invalid_chars:
        file_name = file_name.replace(char, "_")
    return file_name


def select_output_folder():
    folder_selected = filedialog.askdirectory()
    output_folder_entry.delete(0, tk.END)
    output_folder_entry.insert(0, folder_selected)
    save_output_folder(folder_selected)


def send_downloaded_files(
    sender_email, recipient_email, password, subject, body, files
):
    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = recipient_email
    msg["Subject"] = subject

    msg.attach(MIMEText(body, "plain"))

    for file in files:
        with open(file, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())

        encoders.encode_base64(part)

        part.add_header(
            "Content-Disposition",
            f"attachment; filename={os.path.basename(file)}",
        )

        msg.attach(part)

    context = ssl.SSLContext(ssl.PROTOCOL_TLS)
    context.verify_mode = ssl.CERT_REQUIRED
    context.check_hostname = True
    context.load_default_certs()

    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, recipient_email, msg.as_string())


def download_attachments(
    email_address,
    password,
    search_keyword,
    start_date,
    end_date,
    output_folder,
    download_all_attachments,
    file_extension_filter,
    print_output=True,
):
    imap = imapclient.IMAPClient("imap.gmail.com", ssl=True)
    imap.login(email_address, password)
    imap.select_folder("INBOX", readonly=True)
    if not search_keyword.strip():
        return []

    # Remove newline characters from search_keyword
    search_keyword = search_keyword.replace("\n", " ").replace("\r", "")

    date_range = f'SINCE "{start_date.strftime("%d-%b-%Y")}" BEFORE "{(end_date + datetime.timedelta(days=1)).strftime("%d-%b-%Y")}"'
    search_subject = f'SUBJECT "{search_keyword}" {date_range}'
    search_body = f'BODY "{search_keyword}" {date_range}'
    if print_output:
        print(search_subject)
        print(search_body)

    messages_subject = imap.search(search_subject, charset="UTF-8")
    messages_body = imap.search(search_body, charset="UTF-8")
    messages = list(set(messages_subject) | set(messages_body))

    downloaded_files = []

    for msg_id in messages:
        msg_data = imap.fetch([msg_id], ["BODY.PEEK[]", "FLAGS"])[msg_id][b"BODY[]"]
        msg = email.message_from_bytes(msg_data)
        for part in msg.walk():
            if part.get_content_maintype() == "multipart":
                continue

            filename = part.get_filename()
            if not filename or (
                not download_all_attachments
                and not filename.lower().endswith(file_extension_filter.lower())
            ):
                continue

            local_filename = clean_file_name(filename)
            local_filepath = os.path.join(output_folder, local_filename)

            with open(local_filepath, "wb") as f:
                f.write(part.get_payload(decode=True))

            downloaded_files.append(local_filepath)

    imap.logout()

    return downloaded_files


def import_pdf():
    file_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
    if file_path:
        file_name = os.path.basename(file_path)
        files_listbox.insert(tk.END, file_name)
        downloaded_files_list.append(file_path)


def delete_selected_file():
    selected_indices = files_listbox.curselection()
    output_folder = "C:\\Users\\markz\\OneDrive\\Documenten\\downloaded pdfs"  # Adjust the path accordingly

    for index in selected_indices:
        filename = files_listbox.get(index)
        file_path = os.path.join(output_folder, filename)

        # Delete the file from the filesystem
        try:
            os.remove(file_path)
            print(f"Successfully deleted {filename}")
        except OSError as e:
            print(f"{filename} does not exist in the directory. Skipping deletion.")

    # Refresh the listbox to reflect the current state of the output directory
    refresh_files_listbox()


def refresh_files_listbox():
    # Clear the current listbox
    files_listbox.delete(0, tk.END)

    # Fetch all files from the output directory
    output_folder = "C:\\Users\\markz\\OneDrive\\Documenten\\downloaded pdfs"  # Adjust the path accordingly
    files = [
        f
        for f in os.listdir(output_folder)
        if os.path.isfile(os.path.join(output_folder, f))
    ]

    # Populate the listbox with these files
    for file in files:
        files_listbox.insert(tk.END, file)


def download_attachments_gui():
    global downloaded_files_list
    email_address = email_address_entry.get()
    password = password_entry.get()
    search_keyword = search_keyword_entry.get().strip()

    if not search_keyword:
        tk.messagebox.showwarning("Invalid Input", "Please enter a search keyword.")
        return
    save_keyword(search_keyword)

    start_date = start_date_entry.get_date()
    end_date = end_date_entry.get_date()
    output_folder = output_folder_entry.get()
    download_all_attachments = all_attachments_var.get()
    file_extension_filter = file_extension_entry.get()

    downloaded_files_list = download_attachments(
        email_address,
        password,
        search_keyword,
        start_date,
        end_date,
        output_folder,
        download_all_attachments,
        file_extension_filter,
    )

    files_listbox.delete(0, tk.END)
    for file in downloaded_files_list:
        files_listbox.insert(tk.END, os.path.basename(file))

    tk.messagebox.showinfo("Download Complete", "Alle bijlagen zijn gedownload.")


def send_selected_files():
    email_address = email_address_entry.get()
    password = password_entry.get()
    recipient_email = recipient_email_entry.get()

    selected_files_indices = files_listbox.curselection()
    selected_files = [downloaded_files_list[i] for i in selected_files_indices]

    send_downloaded_files(
        email_address,
        recipient_email,
        password,
        "Downloaded Attachments",
        "Here are the downloaded attachments.",
        selected_files,
    )
    tk.messagebox.showinfo(
        "Sending Complete", "De geselecteerde bijlagen zijn verzonden."
    )


def clear_downloaded_files_list():
    global downloaded_files_list
    downloaded_files_list.clear()
    files_listbox.delete(0, tk.END)


def show_files_in_dir():
    global files_listbox
    directory = filedialog.askdirectory()  # Open a dialog to ask for the directory
    if not directory:  # Check if the user selected a directory
        print("No directory selected.")
        return

    # Filter PDF files
    files = [f for f in os.listdir(directory) if f.endswith(".pdf")]

    # Clear previous listbox entries
    files_listbox.delete(0, tk.END)

    for file in files:
        files_listbox.insert(tk.END, file)

    selected_files = []

    def on_ok():
        selected_files = [files_listbox.get(i) for i in files_listbox.curselection()]
        print(selected_files)  # Do something with the selected files

    ok_button = tk.Button(root, text="OK", command=on_ok)
    ok_button.grid(row=10, column=1, padx=5, pady=10)


def parse_pdf_to_excel():
    # get selected files from listbox
    selected_files = files_listbox.curselection()
    if not selected_files:  # Check if any files were selected
        print("No files selected.")
        return

    output_folder = (
        output_folder_entry.get()
    )  # assuming this is where the output folder is stored

    for file_index in selected_files:
        filename = files_listbox.get(file_index)
        file_path = os.path.join(output_folder, filename)  # form the full path

        reader = PdfReader(file_path)
        content = next(iter(reader.pages)).extract_text()

        # Split the content by line breaks and iterate over the lines
        lines = content.split("\n")
        data = []
        for line in lines:
            # Skip lines that contain certain keywords
            if "E-mail" in line or "Rotterdam" in line:
                continue

            # Use regular expressions to extract the fields
            match = re.match(
                r"(\d+)\s+(.*?)(\s{2,})(\d+)\s+([\d,]+)\s+([\d,.%]+)\s+([\d,]+)", line
            )
            if match:  # Check if the line matches the expected format
                # Map the matched groups to their respective column names
                data.append(
                    {
                        "Artikelnr": match.group(1),
                        "Omschrijving/Kleur": match.group(2),
                        "Aantal": match.group(4),
                        "Prijs per stuk": match.group(5).replace(
                            " ", ""
                        ),  # remove any spaces
                        "Korting": match.group(6),
                        "Netto": match.group(7).replace(" ", ""),  # remove any spaces
                    }
                )
            else:
                print(f"Line didn't match the pattern: {line}")

        print(f"Extracted data: {data}")

        # Create a DataFrame from the data
        df = pd.DataFrame(data)

        # Load the existing Excel file or create a new one if it doesn't exist
        try:
            df_existing = pd.read_excel("output.xlsx")
            df = pd.concat([df_existing, df])
        except FileNotFoundError:
            pass

        # Writing DataFrame to Excel
        df.to_excel("output.xlsx", index=False)

        # Load the workbook
        wb = load_workbook(filename="output.xlsx")
        sheet = wb.active

        # Set the width of the columns
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            column_name = column[0].value
            if column_name is not None:
                max_length = len(column_name)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = max_length + 2
            else:
                adjusted_width = 11

            # Make "Omschrijving/Kleur" column 2 times wider
            if column_name == "Omschrijving/Kleur":
                adjusted_width *= 1.5

            sheet.column_dimensions[
                get_column_letter(column[0].column)
            ].width = adjusted_width

        # Save the workbook
        wb.save("output.xlsx")


def clear_output_file():
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    wb.save("output.xlsx")
    print("Output file has been cleared.")


def generate_summary():
    file_counter = collections.Counter()
    total_size = 0

    for file in downloaded_files_list:
        file_extension = os.path.splitext(file)[1]
        file_size = os.path.getsize(file)
        file_counter[file_extension] += 1
        total_size += file_size

    summary_text = f"Total attachments: {len(downloaded_files_list)}\n"
    summary_text += f"Total size: {total_size / 1024:.2f} KB\n"
    summary_text += "\nFile type breakdown:\n"

    for file_ext, count in file_counter.items():
        summary_text += f"{file_ext}: {count}\n"

    tk.messagebox.showinfo("Summary", summary_text)


def download_attachments_all_keywords_gui():
    global downloaded_files_list
    email_address = email_address_entry.get()
    password = password_entry.get()
    start_date = start_date_entry.get_date()
    end_date = end_date_entry.get_date()
    output_folder = output_folder_entry.get()
    download_all_attachments = all_attachments_var.get()
    file_extension_filter = file_extension_entry.get()

    downloaded_files_list = []

    # Create a new top-level window
    progress_window = tk.Toplevel(root)
    progress_window.title("Downloading...")

    # Create a Label
    progress_label = tk.Label(progress_window, text="Downloading attachments...")
    progress_label.pack()

    # Create a Progressbar
    progress = ttk.Progressbar(progress_window, length=300)
    progress.pack()

    # Initialize the progress bar
    progress["maximum"] = len(unique_keywords)
    progress["value"] = 0

    for keyword in unique_keywords:
        current_downloaded_files = download_attachments(
            email_address,
            password,
            keyword,
            start_date,
            end_date,
            output_folder,
            download_all_attachments,
            file_extension_filter,
            print_output=False,
        )

        downloaded_files_list.extend(current_downloaded_files)

        # Update the progress bar
        progress["value"] += 1
        progress_window.update_idletasks()

    files_listbox.delete(0, tk.END)
    for file in sorted(downloaded_files_list, key=os.path.basename):
        files_listbox.insert(tk.END, os.path.basename(file))

    progress_window.destroy()

    tk.messagebox.showinfo("Download Complete", "Alle bijlagen zijn gedownload.")


def open_pdf_file(root, file_path=None):
    global pdf_window
    output_folder = "C:\\Users\\markz\\OneDrive\\Documenten\\downloaded pdfs"  # Set the correct path here

    if pdf_window is not None:
        messagebox.showwarning(
            "Warning",
            "Please close the currently open PDF window before opening another.",
        )
        return

    if not file_path:  # Ask for a file if no path is provided
        file_path = filedialog.askopenfilename(
            initialdir=output_folder, filetypes=[("PDF files", "*.pdf")]
        )

    if file_path:
        images = convert_from_path(file_path, dpi=300)
        if images:
            image = images[0]
            image.thumbnail((500, 500))
            photo = ImageTk.PhotoImage(
                image, master=root
            )  # Pass the root window as the master

            # Create a new Toplevel window
            pdf_window = tk.Toplevel(root)
            pdf_window.title("PDF Preview")

            # Create a label for the PDF preview on the new window
            pdf_preview_label = tk.Label(pdf_window, image=photo)
            pdf_preview_label.image = photo  # Keep a reference to the image object
            pdf_preview_label.pack()

        else:
            messagebox.showerror("Error", "Failed to load the PDF file.")


def on_listbox_double_click(event):
    print("Double click detected")  # Debugging line
    widget = event.widget  # Get the widget that triggered the event
    selection = widget.curselection()
    if selection:
        filename = widget.get(selection[0])
        output_folder = (
            output_folder_entry.get()
        )  # assuming this is where the output folder is stored
        full_path = os.path.join(output_folder, filename)  # form the full path
        open_pdf_file(
            root, full_path
        )  # call open_pdf_file function passing the root and the full path


def close_pdf_file():
    global pdf_window
    if pdf_window is not None:
        pdf_window.destroy()
        pdf_window = None


def save_email_addresses(email_address, recipient_email):
    with open("email_addresses.txt", "w") as file:
        file.write(email_address + "\n")
        file.write(recipient_email + "\n")


def load_email_addresses():
    try:
        with open("email_addresses.txt", "r") as file:
            email_address = file.readline().strip()
            recipient_email = file.readline().strip()
        return email_address, recipient_email
    except FileNotFoundError:
        return "", ""


class ToolTip(object):
    def __init__(self, widget, text="widget info"):
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)
        self.tipwindow = None

    def enter(self, event=None):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(
            tw,
            text=self.text,
            justify=tk.LEFT,
            background="#ffffe0",
            relief=tk.SOLID,
            borderwidth=1,
            font=("tahoma", "8", "normal"),
        )
        label.pack(ipadx=1)

    def leave(self, event=None):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()


root = tk.Tk()
root.title("Python Email Downloader")

# E-mail Address
email_address_label = tk.Label(root, text="E-mailadres:")
email_address_label.grid(row=0, column=0, padx=5, pady=5, sticky="W")
email_address_entry = tk.Entry(root)
email_address_entry.grid(row=0, column=1, padx=5, pady=5, sticky="WE")

# Password
password_label = tk.Label(root, text="Wachtwoord:")
password_label.grid(row=1, column=0, padx=5, pady=5, sticky="W")
password_entry = tk.Entry(root, show="*")
password_entry.grid(row=1, column=1, padx=5, pady=5, sticky="WE")

# Search Keyword
search_keyword_label = tk.Label(root, text="Zoekwoord:")
search_keyword_label.grid(row=2, column=0, padx=5, pady=5, sticky="W")
search_keyword_entry = ttk.Combobox(root)
search_keyword_entry.grid(row=2, column=1, padx=5, pady=5, sticky="WE")
remove_keyword_button = tk.Button(root, text="Verwijderen", command=remove_keyword)
remove_keyword_button.grid(row=2, column=2, padx=5, pady=5)

# Start Date
start_date_label = tk.Label(root, text="Startdatum:")
start_date_label.grid(row=3, column=0, padx=5, pady=5, sticky="W")
start_date_entry = DateEntry(root, date_pattern="dd-mm-yyyy")
start_date_entry.grid(row=3, column=1, padx=5, pady=5, sticky="WE")

# End Date
end_date_label = tk.Label(root, text="Einddatum:")
end_date_label.grid(row=4, column=0, padx=5, pady=5, sticky="W")
end_date_entry = DateEntry(root, date_pattern="dd-mm-yyyy")
end_date_entry.grid(row=4, column=1, padx=5, pady=5, sticky="WE")

# Output Folder
output_folder_label = tk.Label(root, text="Uitvoermap:")
output_folder_label.grid(row=5, column=0, padx=5, pady=5, sticky="W")
output_folder_entry = tk.Entry(root)
output_folder_entry.grid(row=5, column=1, padx=5, pady=5, sticky="WE")
browse_button = tk.Button(root, text="Bladeren", command=select_output_folder)
browse_button.grid(row=5, column=2, padx=5, pady=5)

# All Attachments Checkbox
all_attachments_var = tk.IntVar()
all_attachments_checkbox = tk.Checkbutton(
    root, text="Download alle bijlagen", variable=all_attachments_var
)
all_attachments_checkbox.grid(row=6, column=0, columnspan=1, padx=5, pady=5)

# Download All Attachments
download_all_button = tk.Button(
    root,
    text="Download Alle zoekwoorden tegelijk",
    command=download_attachments_all_keywords_gui,
)
download_all_button.grid(
    row=6, column=2, columnspan=2, pady=(10, 0), padx=5, sticky="w"
)
download_all_button_tooltip = ToolTip(
    download_all_button, "Click to download alle zoekwoorden pdf's tegelijk."
)

# File Extension Filter
file_extension_label = tk.Label(root, text="Bestandsextensie filter:")
file_extension_label.grid(row=7, column=0, padx=5, pady=5, sticky="W")
file_extension_entry = tk.Entry(root)
file_extension_entry.grid(row=7, column=1, padx=5, pady=5, sticky="WE")
file_extension_entry.insert(0, ".pdf")

# Recipient Email
recipient_email_label = tk.Label(root, text="Ontvanger E-mail:")
recipient_email_label.grid(row=8, column=0, padx=5, pady=5, sticky="W")
recipient_email_entry = tk.Entry(root)
recipient_email_entry.grid(row=8, column=1, padx=5, pady=5, sticky="WE")

# Download Button
download_button = tk.Button(root, text="Downloaden", command=download_attachments_gui)
download_button.grid(row=9, column=0, padx=5, pady=10)
download_button_tooltip = ToolTip(download_button, "Click to download.")

# Files Listbox
files_listbox = tk.Listbox(root, selectmode=tk.MULTIPLE)
files_listbox.grid(row=9, column=1, padx=5, pady=10, sticky="WE")

# Send Button
send_button = tk.Button(root, text="Verzenden", command=send_selected_files)
send_button.grid(row=9, column=2, padx=5, pady=10)
send_button_tooltip = ToolTip(send_button, "Click to send selected files.")

# Clear Button
clear_button = tk.Button(root, text="Lijst wissen", command=clear_downloaded_files_list)
clear_button.grid(row=9, column=3, padx=5, pady=10)
clear_button_tooltip = ToolTip(clear_button, "Click to clear the list.")

# Summary Button
summary_button = tk.Button(root, text="Summary", command=generate_summary)
summary_button.grid(row=9, column=4, padx=5, pady=10)
summary_button_tooltip = ToolTip(summary_button, "Click to generate a summary.")

# Close pdf preview Button
close_button = tk.Button(root, text="Close PDF", command=close_pdf_file)
close_button.grid(row=12, column=0, padx=5, pady=5)
close_button_tooltip = ToolTip(close_button, "Click to close the preview pdf.")

# Delete Button
delete_button = tk.Button(root, text="Delete Selected", command=delete_selected_file)
delete_button.grid(row=8, column=2, columnspan=2, pady=(10, 0), padx=10, sticky="WE")
delete_button_tooltip = ToolTip(delete_button, "Click to delete selected files.")

pdf_button = tk.Button(root, text="Open PDF", command=lambda: open_pdf_file(root))
pdf_button.grid(row=11, column=0, padx=5, pady=5)
pdf_button_tooltip = ToolTip(pdf_button, "Click to open a PDF file.")

delete_dir_button = tk.Button(root, text="Delete Directory", command=delete_directory)
delete_dir_button.grid(row=11, column=4, padx=5, pady=5)

# Parse PDF Button
parse_pdf_button = tk.Button(
    root, text="Parse PDF to Excel", command=parse_pdf_to_excel
)
parse_pdf_button.grid(row=10, column=2, padx=5, pady=5)
parse_pdf_button_tooltip = ToolTip(parse_pdf_button, "Click to parse PDF to Excel.")

# Clear Output File Button
clear_output_file_button = tk.Button(
    root, text="Clear Output File", command=clear_output_file
)
clear_output_file_button.grid(row=10, column=3, padx=5, pady=5)

quit_button = tk.Button(root, text="Afsluiten", command=close_application)
quit_button.grid(row=14, column=4, padx=5, pady=10)
quit_button_tooltip = ToolTip(quit_button, "Click to close the application.")

pdf_preview_label = tk.Label(root)
pdf_preview_label.grid(row=11, column=1, padx=5, pady=5, columnspan=4)

import_pdf_button = tk.Button(root, text="Import PDF", command=import_pdf)
import_pdf_button.grid(row=12, column=4, sticky="W", padx=(10, 0), pady=(10, 10))
import_pdf_button_tooltip = ToolTip(import_pdf_button, "Click to import a PDF file.")

email_address, recipient_email = load_email_addresses()
email_address_entry.insert(0, email_address)
recipient_email_entry.insert(0, recipient_email)

stored_keywords = load_keywords()
if stored_keywords:
    search_keyword_entry["values"] = stored_keywords
    search_keyword_entry.set(stored_keywords[-1])

stored_output_folder = load_output_folder()
if stored_output_folder:
    output_folder_entry.insert(0, stored_output_folder)


def save_dates():
    with open("dates_settings.txt", "w") as file:
        file.write(start_date_entry.get() + "\n")
        file.write(end_date_entry.get() + "\n")


def on_closing():
    email_address = email_address_entry.get()
    recipient_email = recipient_email_entry.get()
    save_email_addresses(email_address, recipient_email)
    save_dates()
    root.destroy()


root.protocol("WM_DELETE_WINDOW", on_closing)

# Empty label to fill space
empty_label = tk.Label(root)
empty_label.grid(row=15, sticky="ns")

# Configuring row weight for the empty label to be large to consume remaining space
root.grid_rowconfigure(15, weight=1)

# Scrolling text
scroll_text_var = tk.StringVar()
scroll_label = tk.Label(
    root, textvariable=scroll_text_var, anchor="w", font=("Helvetica", 20)
)


def load_dates():
    try:
        with open("dates_settings.txt", "r") as file:
            start_date = file.readline().strip()
            end_date = file.readline().strip()
        return start_date, end_date
    except FileNotFoundError:
        return None, None


start_date, end_date = load_dates()
if start_date:
    start_date_entry.set_date(start_date)
if end_date:
    end_date_entry.set_date(end_date)


def scroll_text(message, delay=100):
    if len(message) > 0:
        message = message[1:] + message[0]  # Move the first character to the end
        scroll_text_var.set(message)
        root.after(delay, lambda: scroll_text(message, delay))


scroll_text("   Copyright 2023 - Mark Zandbergen.   ")

# Find the maximum row number among all grid slaves
max_row = max(child.grid_info()["row"] for child in root.grid_slaves())
scroll_label.grid(
    row=max_row + 1, column=0, columnspan=5, sticky="WE"
)  # Place the scroll_label below all existing widgets

# Window Configuration
window_width = 680
window_height = 800
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x_coord = int((screen_width / 2) - (window_width / 2))
y_coord = int((screen_height / 2) - (window_height / 2))
root.geometry(f"{window_width}x{window_height}+{x_coord}+{y_coord}")

files_listbox.bind("<Double-Button-1>", on_listbox_double_click)


root.mainloop()
